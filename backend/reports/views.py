"""
reports/views.py

This app has no models of its own — it aggregates data from products,
suppliers, warehouses, stock, and purchase_orders to power:
1. The Dashboard (summary cards + chart data)
2. Stock / Supplier / Warehouse / Purchase Order reports, exportable
   as CSV or Excel.
"""

import csv
from datetime import timedelta

from django.db.models import Count, F, Sum
from django.http import HttpResponse
from django.utils import timezone
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from products.models import Product
from purchase_orders.models import PurchaseOrder
from stock.models import StockItem, StockTransaction
from suppliers.models import Supplier
from warehouses.models import Warehouse

try:
    import openpyxl
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:  # pragma: no cover - openpyxl is in requirements.txt
    OPENPYXL_AVAILABLE = False


class DashboardSummaryView(APIView):
    """
    GET /api/reports/dashboard/summary/
    Powers the 5 summary cards: Total Products, Total Warehouses,
    Total Suppliers, Low Stock Items, Pending Purchase Orders.
    """

    permission_classes = [IsAuthenticated]

    def get(self, request):
        low_stock_count = StockItem.objects.filter(
            current_stock__lte=F("minimum_stock_level")
        ).count()
        pending_po_count = PurchaseOrder.objects.filter(
            status__in=[PurchaseOrder.Status.DRAFT, PurchaseOrder.Status.APPROVED]
        ).count()

        data = {
            "total_products": Product.objects.filter(is_active=True).count(),
            "total_warehouses": Warehouse.objects.count(),
            "total_suppliers": Supplier.objects.filter(status=Supplier.Status.ACTIVE).count(),
            "low_stock_items": low_stock_count,
            "pending_purchase_orders": pending_po_count,
        }
        return Response(data)


class MonthlyStockMovementView(APIView):
    """
    GET /api/reports/dashboard/stock-movement/?months=6
    Returns total quantity of STOCK_IN vs STOCK_OUT per month, for the
    'Monthly Stock Movement' chart (Recharts bar/line chart).
    """

    permission_classes = [IsAuthenticated]

    def get(self, request):
        months = int(request.query_params.get("months", 6))
        since = timezone.now() - timedelta(days=30 * months)

        qs = (
            StockTransaction.objects.filter(created_at__gte=since)
            .filter(transaction_type__in=[
                StockTransaction.TransactionType.STOCK_IN,
                StockTransaction.TransactionType.STOCK_OUT,
            ])
        )

        # Group manually by (year, month, type) since cross-database
        # TruncMonth behavior can vary; this keeps it simple & portable.
        buckets = {}
        for txn in qs.only("created_at", "transaction_type", "quantity"):
            key = txn.created_at.strftime("%Y-%m")
            buckets.setdefault(key, {"month": key, "stock_in": 0, "stock_out": 0})
            if txn.transaction_type == StockTransaction.TransactionType.STOCK_IN:
                buckets[key]["stock_in"] += txn.quantity
            else:
                buckets[key]["stock_out"] += txn.quantity

        result = sorted(buckets.values(), key=lambda r: r["month"])
        return Response(result)


class PurchaseTrendsView(APIView):
    """
    GET /api/reports/dashboard/purchase-trends/?months=6
    Returns count + total value of purchase orders created per month,
    for the 'Purchase Trends' chart.
    """

    permission_classes = [IsAuthenticated]

    def get(self, request):
        months = int(request.query_params.get("months", 6))
        since = timezone.now() - timedelta(days=30 * months)

        pos = PurchaseOrder.objects.filter(created_at__gte=since).prefetch_related("lines")
        buckets = {}
        for po in pos:
            key = po.created_at.strftime("%Y-%m")
            buckets.setdefault(key, {"month": key, "po_count": 0, "total_value": 0})
            buckets[key]["po_count"] += 1
            buckets[key]["total_value"] += float(po.total_amount)

        result = sorted(buckets.values(), key=lambda r: r["month"])
        return Response(result)


class InventoryValueView(APIView):
    """
    GET /api/reports/dashboard/inventory-value/
    Returns total inventory value (current_stock * cost_price) grouped
    by warehouse, for the 'Inventory Value' chart.
    """

    permission_classes = [IsAuthenticated]

    def get(self, request):
        rows = (
            StockItem.objects.select_related("warehouse", "product")
            .values("warehouse__name")
            .annotate(
                total_value=Sum(F("current_stock") * F("product__cost_price")),
                total_units=Sum("current_stock"),
            )
            .order_by("warehouse__name")
        )
        result = [
            {
                "warehouse": row["warehouse__name"],
                "total_value": float(row["total_value"] or 0),
                "total_units": row["total_units"] or 0,
            }
            for row in rows
        ]
        return Response(result)


# ---------------------------------------------------------------------
# Report generation + export (CSV / Excel)
# ---------------------------------------------------------------------

class BaseReportView(APIView):
    """
    Shared logic for all report endpoints: subclasses implement
    get_rows()/get_headers(), and this base class handles rendering
    that data as JSON (default), CSV (?format=csv), or Excel
    (?format=excel).
    """

    permission_classes = [IsAuthenticated]
    report_filename = "report"

    def get_headers(self):
        raise NotImplementedError

    def get_rows(self, request):
        raise NotImplementedError

    def get(self, request):
        headers = self.get_headers()
        rows = self.get_rows(request)
        export_format = request.query_params.get("format", "json")

        if export_format == "csv":
            return self._render_csv(headers, rows)
        if export_format == "excel":
            return self._render_excel(headers, rows)
        return Response([dict(zip(headers, row)) for row in rows])

    def _render_csv(self, headers, rows):
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="{self.report_filename}.csv"'
        writer = csv.writer(response)
        writer.writerow(headers)
        writer.writerows(rows)
        return response

    def _render_excel(self, headers, rows):
        if not OPENPYXL_AVAILABLE:
            return Response(
                {"detail": "Excel export requires the 'openpyxl' package on the server."},
                status=500,
            )
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.report_filename[:31]
        ws.append(headers)
        for row in rows:
            ws.append(list(row))
        # Auto-fit column widths roughly, based on content length.
        for i, header in enumerate(headers, start=1):
            col_letter = get_column_letter(i)
            max_len = max([len(str(header))] + [len(str(r[i - 1])) for r in rows]) if rows else len(str(header))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="{self.report_filename}.xlsx"'
        wb.save(response)
        return response


class StockReportView(BaseReportView):
    """GET /api/reports/stock/?format=csv|excel"""

    report_filename = "stock_report"

    def get_headers(self):
        return ["SKU", "Product", "Warehouse", "Current Stock", "Reserved", "Available", "Min Level", "Low Stock"]

    def get_rows(self, request):
        items = StockItem.objects.select_related("product", "warehouse").all()
        warehouse_id = request.query_params.get("warehouse")
        if warehouse_id:
            items = items.filter(warehouse_id=warehouse_id)
        return [
            [
                item.product.sku,
                item.product.name,
                item.warehouse.name,
                item.current_stock,
                item.reserved_stock,
                item.available_stock,
                item.minimum_stock_level,
                "YES" if item.is_low_stock else "NO",
            ]
            for item in items
        ]


class SupplierReportView(BaseReportView):
    """GET /api/reports/suppliers/?format=csv|excel"""

    report_filename = "supplier_report"

    def get_headers(self):
        return ["Name", "Contact Person", "Email", "Phone", "Status", "Total Products Supplied"]

    def get_rows(self, request):
        suppliers = Supplier.objects.annotate(product_count=Count("products"))
        return [
            [s.name, s.contact_person or "", s.email, s.phone, s.status, s.product_count]
            for s in suppliers
        ]


class WarehouseReportView(BaseReportView):
    """GET /api/reports/warehouses/?format=csv|excel"""

    report_filename = "warehouse_report"

    def get_headers(self):
        return ["Name", "Location", "Capacity", "Manager", "Current Utilization", "Utilization %"]

    def get_rows(self, request):
        warehouses = Warehouse.objects.select_related("manager").all()
        rows = []
        for w in warehouses:
            utilization = w.current_utilization
            pct = round((utilization / w.capacity) * 100, 2) if w.capacity else 0
            rows.append([
                w.name,
                w.location,
                w.capacity,
                w.manager.get_full_name() if w.manager else "",
                utilization,
                pct,
            ])
        return rows


class PurchaseOrderReportView(BaseReportView):
    """GET /api/reports/purchase-orders/?format=csv|excel&status=APPROVED"""

    report_filename = "purchase_order_report"

    def get_headers(self):
        return ["PO Number", "Supplier", "Warehouse", "Date", "Status", "Total Amount"]

    def get_rows(self, request):
        pos = PurchaseOrder.objects.select_related("supplier", "warehouse").prefetch_related("lines")
        status_filter = request.query_params.get("status")
        if status_filter:
            pos = pos.filter(status=status_filter)
        return [
            [po.po_number, po.supplier.name, po.warehouse.name, po.date.isoformat(), po.status, float(po.total_amount)]
            for po in pos
        ]
